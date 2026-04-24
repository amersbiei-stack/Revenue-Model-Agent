"""Exhaustive verification of FY26 Plan row 16 and surrounding context."""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

path = Path(r"C:\Users\amers\Downloads\Revenue Model\Revenue Model - 04062026 (Internal).xlsm")
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
fp = wb['FY26 Plan']

print(f"FY26 Plan dimensions: {fp.dimensions}")
print(f"Max row: {fp.max_row}, Max col: {fp.max_column}\n")

# 1. Merged cells — openpyxl stores value in top-left only, rest read as None
print("=== Merged cell ranges in FY26 Plan ===")
merges = list(fp.merged_cells.ranges)
print(f"Total merged ranges: {len(merges)}")
for mr in merges[:30]:
    print(f"  {mr}")
if len(merges) > 30:
    print(f"  ... and {len(merges) - 30} more")

# 2. Dump row 16 fully, every column, raw values
print("\n=== Row 16 — every column, raw cell.value ===")
for c in range(1, 45):
    v = fp.cell(row=16, column=c).value
    letter = get_column_letter(c)
    if v is not None:
        print(f"  {letter}16 (col {c}): {v!r}")
    else:
        # check if this cell is inside a merge
        for mr in merges:
            if mr.min_row <= 16 <= mr.max_row and mr.min_col <= c <= mr.max_col:
                anchor = fp.cell(row=mr.min_row, column=mr.min_col).value
                print(f"  {letter}16 (col {c}): [MERGED with {mr}, anchor={anchor!r}]")
                break

# 3. Also dump row 14 (the Q1-2026 / Q2-2026 header row) — same treatment
print("\n=== Row 14 — every column, raw cell.value (merges resolved) ===")
for c in range(1, 45):
    v = fp.cell(row=14, column=c).value
    letter = get_column_letter(c)
    if v is not None:
        print(f"  {letter}14 (col {c}): {v!r}")
    else:
        for mr in merges:
            if mr.min_row <= 14 <= mr.max_row and mr.min_col <= c <= mr.max_col:
                anchor = fp.cell(row=mr.min_row, column=mr.min_col).value
                if anchor is not None:
                    print(f"  {letter}14 (col {c}): [MERGED with {mr}, anchor={anchor!r}]")
                break

# 4. Spot-check: is the value in W16 literally "External"?
print("\n=== Spot check W16 and nearby ===")
for addr in ['V16', 'W16', 'X16', 'Y16', 'Z16', 'AA16']:
    cell = fp[addr]
    print(f"  {addr}: value={cell.value!r}")

# 5. Any *other* row that might label these blocks? Row 15 or row 17 labels?
print("\n=== Row 15 — any content? ===")
for c in range(1, 45):
    v = fp.cell(row=15, column=c).value
    if v is not None:
        print(f"  {get_column_letter(c)}15: {v!r}")

# 6. Finally: is there a "Q1-2026" / "Q2-2026" pattern in ROW 14 that extends to W column?
# This confirms whether the second block (cols W+) even has a quarter label, or if it's something else entirely.
print("\n=== FY26 Plan — visual sanity: col A..AH of rows 14, 15, 16, 17 side by side ===")
header_cols = list(range(2, 35))  # B through AH
print("row | " + " | ".join(f"{get_column_letter(c):>5}" for c in header_cols))
for r in [14, 15, 16, 17]:
    vals = []
    for c in header_cols:
        v = fp.cell(row=r, column=c).value
        if v is None:
            # resolve merge if any
            for mr in merges:
                if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                    v = fp.cell(row=mr.min_row, column=mr.min_col).value
                    break
        s = "" if v is None else str(v)
        vals.append(f"{s[:5]:>5}")
    print(f"{r:>3} | " + " | ".join(vals))
