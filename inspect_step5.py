"""Inspect the 5 Step 5 tabs: header row, data range, total cell."""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

path = Path(r"C:\Users\amers\Downloads\Revenue Model\Revenue Model - 04062026 (Internal).xlsm")
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)

TABS = ['Units Bookings DV', '$$ Bookings DV', 'Sub Europe Units', 'Migration Units', 'AS']

def find_month_header_row(ws, sample_label='2026M04', max_row=30, max_col=250):
    """Find the row that contains the month label (e.g., 2026M04)."""
    hits = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip()
            if s == sample_label:
                hits.append((r, c, get_column_letter(c)))
    return hits

def scan_col_a_b(ws, max_row=1000):
    """Walk column A and B, return list of (row, a_val, b_val) where anything non-empty."""
    out = []
    for r in range(1, max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if (a is not None and str(a).strip() != "") or (b is not None and str(b).strip() != ""):
            out.append((r, a, b))
    return out

for tab_name in TABS:
    if tab_name not in wb.sheetnames:
        print(f"\n=== {tab_name} ===  NOT FOUND")
        continue
    ws = wb[tab_name]
    print(f"\n================ {tab_name} ================")
    print(f"Dimensions: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")

    # 1. Find month header row by locating '2026M04'
    hits = find_month_header_row(ws, '2026M04')
    if hits:
        print(f"Found '2026M04' at: {hits[:3]}{'...' if len(hits) > 3 else ''}")
    else:
        # Try other month labels
        for lbl in ['2026M03', '2026M01', '2025M12', '2026M05']:
            hits = find_month_header_row(ws, lbl)
            if hits:
                print(f"Found '{lbl}' at: {hits[:3]}")
                break
        else:
            print("No YYYYMmm month labels found in first 30 rows.")

    # 2. Column A/B labels — compressed view, show transitions
    print(f"\n  Column A/B labels (rows where non-empty):")
    items = scan_col_a_b(ws, max_row=300)

    # Collapse runs where column B has repeating values
    if items:
        prev_b = None
        start = None
        for (r, a, b) in items:
            bkey = str(b) if b is not None else ""
            akey = str(a)[:30] if a is not None else ""
            if bkey != prev_b or akey:
                if start is not None and prev_b is not None:
                    pass
                start = r
                prev_b = bkey
                ashort = akey[:30] if akey else ""
                bshort = bkey[:40] if bkey else ""
                print(f"    r{r:>4}: A={ashort!r:<32} B={bshort!r}")
        # show total count of rows with any content
        print(f"  Total rows with A/B content: {len(items)}")
    else:
        print("  (column A and B both empty for rows 1..300)")
