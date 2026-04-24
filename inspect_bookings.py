"""Read Summary By Region formulas at rows 22, 29, 36, 64, 85, 92 and col A labels."""
import openpyxl
from pathlib import Path
import re

path = Path(r"C:\Users\amers\Downloads\Revenue Model\Revenue Model - 04062026 (Internal).xlsm")
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
sm = wb['Summary By Region']
bk = wb['$$ Bookings DV']

print("=== Summary By Region — col B labels, rows 1..100 ===\n")
for r in range(1, 101):
    a = sm.cell(row=r, column=1).value
    b = sm.cell(row=r, column=2).value
    c = sm.cell(row=r, column=3).value
    d = sm.cell(row=r, column=4).value
    any_label = any(v not in (None, "") for v in (a, b, c, d))
    if any_label:
        def fmt(v):
            if v is None: return ""
            s = str(v)
            return s[:35] + ("…" if len(s) > 35 else "")
        print(f"  r{r:>3}: A={fmt(a)!r:<25} B={fmt(b)!r:<38} C={fmt(c)!r:<25} D={fmt(d)!r}")

print("\n\n=== Summary By Region — formulas at spec rows (22, 29, 36, 64, 85, 92) ===")
print("Check formulas in a few columns: F (6), G (7), and the last real month col\n")

# Row 4 holds the month headers in this workbook based on the macro — check that
print("Row 4 sample (month headers): col 60..80:")
for col in range(60, 81):
    v = sm.cell(row=4, column=col).value
    if v is not None:
        print(f"  col {col} ({openpyxl.utils.get_column_letter(col)}): {v!r}")

print()
rows_to_check = [22, 29, 36, 64, 85, 92, 165]
for rr in rows_to_check:
    label_b = sm.cell(row=rr, column=2).value
    label_a = sm.cell(row=rr, column=1).value
    print(f"\n--- Summary By Region row {rr}  |  A={label_a!r}  B={label_b!r} ---")
    # Look at columns 5..10 and 60..70 for formulas
    for col in list(range(5, 11)) + list(range(60, 71)):
        v = sm.cell(row=rr, column=col).value
        if v is None:
            continue
        letter = openpyxl.utils.get_column_letter(col)
        s = str(v)
        if len(s) > 140:
            s = s[:140] + "…"
        print(f"    {letter}{rr}: {s}")
