"""FY26 Plan — find ALL month labels in row 2 (dict overwrites hid duplicates)."""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

path = Path(r"C:\Users\amers\Downloads\Revenue Model\Revenue Model - 04062026 (Internal).xlsm")
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
fp = wb['FY26 Plan']

print("=== FY26 Plan — full scan of row 2 (all columns with month labels) ===")
for c in range(1, 50):
    v = fp.cell(row=2, column=c).value
    if v is not None and str(v).strip() != "":
        letter = get_column_letter(c)
        print(f"  col {c:>3} ({letter}): {v!r}")

# Dump rows 1..20 of cols A..AH to see what lives where
print("\n=== FY26 Plan — rows 1..20, col B..AH, showing what's non-empty ===\n")
print(f"{'row':<5} | {'B':<20} | {'C=1Jan':<12} | {'F=M04':<12} | {'V(22)':<12} | {'W=M01?':<12} | {'Z=M04?':<12}")
print("-" * 95)
for r in range(1, 25):
    def c(col):
        v = fp.cell(row=r, column=col).value
        if v is None:
            return ""
        s = str(v)
        return s[:10] + ("…" if len(s) > 10 else "")
    print(f"{r:<5} | {c(2):<20} | {c(3):<12} | {c(6):<12} | {c(22):<12} | {c(23):<12} | {c(26):<12}")

# Check rows 20..90 too — where does plan data actually live?
print("\n=== FY26 Plan — rows 20..92 col B (labels) + col C (first month) + col W (first month of 2nd block) ===")
for r in range(20, 93):
    b = fp.cell(row=r, column=2).value
    c_val = fp.cell(row=r, column=3).value
    w_val = fp.cell(row=r, column=23).value
    if any(v not in (None, "") for v in (b, c_val, w_val)):
        def fmt(v):
            if v is None: return ""
            s = str(v)
            return s[:22] + ("…" if len(s) > 22 else "")
        print(f"  r{r:>3}: B={fmt(b)!r:<24} C={fmt(c_val)!r:<24} W={fmt(w_val)!r}")
